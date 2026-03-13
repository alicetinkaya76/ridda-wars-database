#!/usr/bin/env python3
"""
RIDDA WARS — UNIFIED FIGURE & TABLE GENERATOR v3.0
====================================================
Single script for ALL publication figures and tables.
Target: MDPI Religions Special Issue

Generates:
  FIGURES (11 + 1 supplementary):
    Fig 1:  GenIE Pipeline Architecture (schematic)
    Fig 2:  Source Comparison by Region (grouped horizontal bar)
    Fig 3:  Incorporation Modes by Region (heatmap)
    Fig 4:  Temporal Distribution Year 11 vs 12 AH (grouped bar)
    Fig 5:  Commander Campaign Activity (stacked horizontal bar)
    Fig 6:  Confidence Score Distribution (histogram)
    Fig 7:  Rebellion Cause Taxonomy (2-panel: bar + grouped bar)
    Fig 8:  Source Bias Statistical Test (2-panel: proportional + residuals)
    Fig 9:  Evidence Term Heatmap (term × mode)
    Fig 10: Cross-Model Agreement Heatmap (4×4 kappa matrix)
    Fig S1: Overall Mode Pie Chart (supplementary / graphical abstract)

  TABLES (9 CSV):
    Table 1: Campaign Phases
    Table 2: False Prophets
    Table 3: Arabic Indicators for Incorporation Modes
    Table 4: Dataset Summary Statistics
    Table 5: Top Commanders and Campaign Regions
    Table 6: Cause × Mode Contingency Table
    Table 7: Validation Summary
    Table 8: Cross-Model Agreement — Mode
    Table 9: Cross-Model Agreement — Cause

  OTHER:
    annotation_sheet_IAA_v2.xlsx — Inter-annotator agreement sheet
    source_bias_stats.json — Statistical test results
    ridda_enriched_with_causes.json — Dataset with cause taxonomy

Usage:
    python ridda_figures_unified.py --data-dir data --output-dir figures
    python ridda_figures_unified.py --figures 1 2 3     # specific figures
    python ridda_figures_unified.py --tables-only        # tables only
    python ridda_figures_unified.py --no-annotation       # skip Excel generation

Authors: Gökalp & Çetinkaya (Selçuk University)
"""

import json
import csv
import os
import sys
import random
import argparse
from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime

import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyBboxPatch, Rectangle
from matplotlib.lines import Line2D
import matplotlib.gridspec as gridspec
from matplotlib.colors import LinearSegmentedColormap, Normalize
import matplotlib.ticker as ticker

from scipy import stats as scipy_stats

# ============================================================================
# MDPI PUBLICATION STYLE
# ============================================================================

PALETTE = {
    # Incorporation modes
    'subjugation': '#B71C1C',   # Deep crimson
    'submission':  '#1B5E20',   # Forest green
    'mixed':       '#E65100',   # Burnt orange
    # Sources
    'tabari':      '#0D47A1',   # Navy blue
    'baladhuri':   '#4A148C',   # Deep purple
    # Rebellion causes
    'false_prophet':    '#6A1B9A',
    'zakat_refusal':    '#00838F',
    'tribal_autonomy':  '#E65100',
    'mixed_political':  '#546E7A',
    # General
    'primary':     '#212121',
    'secondary':   '#616161',
    'grid':        '#E0E0E0',
    'bg_light':    '#FAFAFA',
    'accent':      '#00695C',
}

MODE_ORDER = ['SUBJUGATION', 'MIXED', 'SUBMISSION']
MODE_COLORS = {m: PALETTE[m.lower()] for m in MODE_ORDER}
MODE_LABELS = {
    'SUBJUGATION': 'Subjugation (qitāl)',
    'MIXED': 'Mixed',
    'SUBMISSION': 'Submission (ṭāʿa)',
}

CAUSE_ORDER = ['false_prophet', 'zakat_refusal', 'tribal_autonomy', 'mixed_political']
CAUSE_LABELS = {
    'false_prophet': 'False Prophet\nMovement',
    'zakat_refusal': 'Zakāt Refusal',
    'tribal_autonomy': 'Tribal\nAutonomy',
    'mixed_political': 'Mixed /\nUnclear',
}
CAUSE_LABELS_SHORT = {
    'false_prophet': 'False Prophet',
    'zakat_refusal': 'Zakāt Refusal',
    'tribal_autonomy': 'Tribal Autonomy',
    'mixed_political': 'Mixed/Unclear',
}
CAUSE_COLORS = {c: PALETTE[c] for c in CAUSE_ORDER}

# MDPI specs
SINGLE_COL = 3.5   # inches (89 mm)
DOUBLE_COL = 7.0   # inches (178 mm)
DPI = 300

plt.rcParams.update({
    'figure.facecolor': 'white', 'axes.facecolor': 'white',
    'savefig.facecolor': 'white', 'figure.dpi': DPI, 'savefig.dpi': DPI,
    'font.family': 'serif',
    'font.serif': ['DejaVu Serif', 'Times New Roman', 'Georgia'],
    'font.size': 8, 'axes.titlesize': 9, 'axes.titleweight': 'bold',
    'axes.labelsize': 8, 'xtick.labelsize': 7, 'ytick.labelsize': 7,
    'legend.fontsize': 7, 'legend.framealpha': 0.95,
    'axes.linewidth': 0.6, 'axes.grid': False,
    'lines.linewidth': 1.0, 'patch.linewidth': 0.4,
    'axes.spines.top': False, 'axes.spines.right': False,
})


# ============================================================================
# DATA LOADING & UTILITIES
# ============================================================================

def load_data(data_dir):
    with open(os.path.join(data_dir, 'ridda_combined_enriched.json')) as f:
        enriched = json.load(f)
    with open(os.path.join(data_dir, 'ridda_combined_scholarly.json')) as f:
        scholarly = json.load(f)
    return enriched, scholarly

def get_region_short(event):
    r = event.get('_region_en') or event.get('region_english', '')
    return {'Najd (Central Arabia)': 'Najd', 'al-Yaman (Yemen)': 'Yemen',
            'ʿUmān (Oman)': 'Oman'}.get(r, r)

def save_fig(fig, output_dir, name):
    path = os.path.join(output_dir, name)
    fig.savefig(path, dpi=DPI, bbox_inches='tight')
    plt.close(fig)
    print(f"  ✓ {name}")
    return path


# ============================================================================
# CAUSE TAXONOMY ENGINE
# ============================================================================

FALSE_PROPHET_IND = {
    'ar': ['مسيلمة','طليحة','سجاح','الأسود العنسي','عبهلة','لقيط بن مالك',
           'الكذاب','المتنبئ','تنبأ','ادعى النبوة','نبي','وحي','رحمان اليمامة'],
    'en': ['musaylima','tulayha','sajah','aswad','laqit','false prophet',
           'prophet','kadhdhab','liar','prophethood','revelation'],
}
ZAKAT_IND = {
    'ar': ['الزكاة','زكاة','منع الزكاة','منعوا الزكاة','الصدقة','صدقات',
           'أدى الزكاة','أبوا أن يؤدوا','حبسوا الصدقة'],
    'en': ['zakat','zakāt','tax','tribute','alms','sadaqa','ṣadaqa',
           'refused to pay','withheld'],
}
AUTONOMY_IND = {
    'ar': ['ارتد','ردة','مرتد','خلع الطاعة','نقض البيعة','استقل','انفرد','خرج عن'],
    'en': ['apostat','autonomy','independence','broke away','renounced',
           'withdrew allegiance','ridda','separated','defected'],
}

def classify_ridda_cause(event):
    texts_ar, texts_en = [], []
    evidence = event.get('evidence', [])
    if isinstance(evidence, list): texts_ar.extend(evidence)
    else: texts_ar.append(str(evidence))
    for f in ['rebel_leader_arabic', 'notes', 'tribe_arabic']:
        texts_ar.append(event.get(f, '') or '')
    for f in ['rebel_leader_english', 'notes', 'tribe_english']:
        texts_en.append(event.get(f, '') or '')
    all_ar = ' '.join(texts_ar)
    all_en = ' '.join(texts_en).lower()

    scores = {'false_prophet': 0, 'zakat_refusal': 0, 'tribal_autonomy': 0}
    for t in FALSE_PROPHET_IND['ar']:
        if t in all_ar: scores['false_prophet'] += 2
    for t in FALSE_PROPHET_IND['en']:
        if t in all_en: scores['false_prophet'] += 1
    for t in ZAKAT_IND['ar']:
        if t in all_ar: scores['zakat_refusal'] += 2
    for t in ZAKAT_IND['en']:
        if t in all_en: scores['zakat_refusal'] += 1
    for t in AUTONOMY_IND['ar']:
        if t in all_ar: scores['tribal_autonomy'] += 2
    for t in AUTONOMY_IND['en']:
        if t in all_en: scores['tribal_autonomy'] += 1

    max_s = max(scores.values())
    if max_s == 0: return 'mixed_political', 0.5
    high = [c for c, s in scores.items() if s >= max_s * 0.7 and s > 0]
    if len(high) > 1 and max_s < 4: return 'mixed_political', 0.6
    return max(scores, key=scores.get), min(1.0, 0.6 + max_s * 0.1)

def add_cause_taxonomy(events):
    for e in events:
        cause, conf = classify_ridda_cause(e)
        e['_ridda_cause'] = cause
        e['_cause_confidence'] = conf
    return events


# ============================================================================
# FIGURES
# ============================================================================

def fig01_pipeline(output_dir):
    """Fig 1: GenIE Pipeline Architecture (schematic)."""
    fig, ax = plt.subplots(figsize=(DOUBLE_COL, 2.8))
    ax.set_xlim(0, 100); ax.set_ylim(0, 30); ax.axis('off')

    box_src = dict(boxstyle='round,pad=0.6', edgecolor=PALETTE['tabari'], linewidth=0.8, facecolor='#E3F2FD')
    box_std = dict(boxstyle='round,pad=0.6', edgecolor='#37474F', linewidth=0.8, facecolor='#ECEFF1')
    box_out = dict(boxstyle='round,pad=0.6', edgecolor=PALETTE['accent'], linewidth=0.8, facecolor='#E0F2F1')

    steps = [(8,15,'OpenITI\nCorpus',box_src),(24,15,'Chunking\n(~20K chars)',box_std),
             (40,15,'Ridda Term\nDetection',box_std),(56,15,'LLM Schema\nExtraction',box_std),
             (72,15,'Enrichment\n& Geocoding',box_std),(88,15,'Ridda Wars\nDatabase',box_out)]
    for x,y,t,s in steps:
        ax.text(x,y,t,ha='center',va='center',fontsize=7.5,fontweight='bold',color=PALETTE['primary'],bbox=s)
    arrow = dict(arrowstyle='->',color='#546E7A',lw=1.2,connectionstyle='arc3,rad=0')
    for i in range(len(steps)-1):
        ax.annotate('',xy=(steps[i+1][0]-6,15),xytext=(steps[i][0]+6,15),arrowprops=arrow)
    subs = [(8,7.5,'al-Ṭabarī\nal-Balādhurī'),(40,7.5,'ridda, musaylima\nmanʿ al-zakāt'),
            (56,7.5,'Claude API\nJSON schema'),(72,7.5,'al-Turayyā\ncoordinates'),(88,7.5,'91 events\n48 tribes')]
    for x,y,t in subs:
        ax.text(x,y,t,ha='center',va='center',fontsize=6,color=PALETTE['secondary'],style='italic')
    ax.text(50,27,'Generative Information Extraction (GenIE) Pipeline',ha='center',fontsize=9,fontweight='bold')
    plt.tight_layout()
    return save_fig(fig, output_dir, 'fig01_pipeline.png')

def fig02_source_comparison(events, output_dir):
    """Fig 2: Source Comparison by Region (grouped horizontal bar)."""
    rs = defaultdict(lambda: {'tabari':0,'baladhuri':0})
    for e in events:
        r = get_region_short(e); s = e.get('_source','')
        if s in ('tabari','baladhuri'): rs[r][s] += 1
    regions = sorted(rs.keys(), key=lambda r: sum(rs[r].values()))
    fig, ax = plt.subplots(figsize=(DOUBLE_COL, 4.0))
    y = np.arange(len(regions)); h = 0.35
    b1 = ax.barh(y+h/2, [rs[r]['tabari'] for r in regions], h, label='al-Ṭabarī (d. 310 AH)', color=PALETTE['tabari'], edgecolor='white', linewidth=0.3)
    b2 = ax.barh(y-h/2, [rs[r]['baladhuri'] for r in regions], h, label='al-Balādhurī (d. 279 AH)', color=PALETTE['baladhuri'], edgecolor='white', linewidth=0.3)
    for bars in [b1,b2]:
        for bar in bars:
            w = bar.get_width()
            if w > 0: ax.text(w+0.3, bar.get_y()+bar.get_height()/2, str(int(w)), va='center', fontsize=6.5, color=PALETTE['secondary'])
    ax.set_yticks(y); ax.set_yticklabels(regions); ax.set_xlabel('Number of Extracted Events')
    ax.legend(loc='lower right', frameon=True, edgecolor=PALETTE['grid'])
    ax.xaxis.set_major_locator(ticker.MaxNLocator(integer=True)); plt.tight_layout()
    return save_fig(fig, output_dir, 'fig02_source_comparison.png')

def fig03_mode_heatmap(events, output_dir):
    """Fig 3: Incorporation Modes by Region (heatmap)."""
    rm = defaultdict(Counter)
    for e in events:
        r = get_region_short(e); m = e.get('incorporation_mode','')
        if m in MODE_ORDER: rm[r][m] += 1
    top = sorted(rm.keys(), key=lambda r: sum(rm[r].values()), reverse=True)[:10]
    mx = np.zeros((len(top), len(MODE_ORDER)))
    for i,r in enumerate(top):
        for j,m in enumerate(MODE_ORDER): mx[i,j] = rm[r].get(m,0)
    fig, ax = plt.subplots(figsize=(SINGLE_COL+1.5, 4.0))
    cmap = LinearSegmentedColormap.from_list('ridda', ['#FFFFFF','#E0F2F1','#80CBC4','#00897B','#004D40'])
    im = ax.imshow(mx, cmap=cmap, aspect='auto', interpolation='nearest')
    ax.set_xticks(range(len(MODE_ORDER))); ax.set_xticklabels([MODE_LABELS[m] for m in MODE_ORDER], rotation=30, ha='right')
    ax.set_yticks(range(len(top))); ax.set_yticklabels(top)
    for i in range(len(top)):
        for j in range(len(MODE_ORDER)):
            v = int(mx[i,j])
            if v > 0: ax.text(j,i,str(v),ha='center',va='center',fontsize=8,fontweight='bold',color='white' if v>=6 else PALETTE['primary'])
    cbar = fig.colorbar(im, ax=ax, shrink=0.8, pad=0.02); cbar.set_label('Event Count', fontsize=7)
    plt.tight_layout()
    return save_fig(fig, output_dir, 'fig03_mode_heatmap.png')

def fig04_temporal(events, output_dir):
    """Fig 4: Temporal Distribution (Year 11 vs 12 AH by Mode)."""
    ym = defaultdict(Counter)
    for e in events:
        y = e.get('year_ah',0); m = e.get('incorporation_mode','')
        if m in MODE_ORDER and y in (11,12): ym[y][m] += 1
    fig, ax = plt.subplots(figsize=(SINGLE_COL+0.5, 3.2))
    x = np.arange(2); w = 0.22
    for i,mode in enumerate(MODE_ORDER):
        vals = [ym[y][mode] for y in [11,12]]
        bars = ax.bar(x+(i-1)*w, vals, w, label=MODE_LABELS[mode], color=MODE_COLORS[mode], edgecolor='white', linewidth=0.3)
        for bar in bars:
            h = bar.get_height()
            if h > 0: ax.text(bar.get_x()+bar.get_width()/2, h+0.5, str(int(h)), ha='center', fontsize=7, color=PALETTE['secondary'])
    ax.set_xticks(x); ax.set_xticklabels(['Year 11 AH\n(632 CE)','Year 12 AH\n(633 CE)'])
    ax.set_ylabel('Number of Events'); ax.legend(loc='upper right', frameon=True, edgecolor=PALETTE['grid'])
    ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True)); plt.tight_layout()
    return save_fig(fig, output_dir, 'fig04_temporal.png')

def fig05_commanders(events, output_dir):
    """Fig 5: Commander Activity (horizontal stacked bar)."""
    cm = defaultdict(Counter)
    for e in events:
        c = e.get('_commander_normalized') or e.get('commander_english','')
        if not c or c in ('None','not specified'): c = '(unspecified)'
        m = e.get('incorporation_mode','')
        if m in MODE_ORDER: cm[c][m] += 1
    top = sorted(cm.keys(), key=lambda c: sum(cm[c].values()), reverse=True)[:8]
    top.reverse()
    fig, ax = plt.subplots(figsize=(DOUBLE_COL, 3.5)); y = np.arange(len(top)); left = np.zeros(len(top))
    for mode in MODE_ORDER:
        vals = np.array([cm[c][mode] for c in top], dtype=float)
        ax.barh(y, vals, left=left, height=0.6, label=MODE_LABELS[mode], color=MODE_COLORS[mode], edgecolor='white', linewidth=0.3)
        for i,(v,l) in enumerate(zip(vals,left)):
            if v >= 2: ax.text(l+v/2, i, str(int(v)), ha='center', va='center', fontsize=6.5, color='white', fontweight='bold')
        left += vals
    for i,c in enumerate(top): ax.text(left[i]+0.5, i, f'n={sum(cm[c].values())}', va='center', fontsize=6.5, color=PALETTE['secondary'])
    ax.set_yticks(y); ax.set_yticklabels(top, fontsize=7); ax.set_xlabel('Number of Events')
    ax.legend(loc='lower right', frameon=True, edgecolor=PALETTE['grid'], ncol=3, fontsize=6.5)
    ax.xaxis.set_major_locator(ticker.MaxNLocator(integer=True)); plt.tight_layout()
    return save_fig(fig, output_dir, 'fig05_commanders.png')

def fig06_confidence(events, output_dir):
    """Fig 6: Confidence Score Distribution (histogram)."""
    confs = [e.get('confidence',0) for e in events]
    fig, ax = plt.subplots(figsize=(SINGLE_COL+0.5, 3.0))
    bins = np.arange(0.65, 1.05, 0.05)
    n, bo, patches = ax.hist(confs, bins=bins, color=PALETTE['accent'], alpha=0.85, edgecolor='white', linewidth=0.5)
    for p, le in zip(patches, bo[:-1]):
        if le < 0.8: p.set_facecolor('#EF5350')
        elif le < 0.9: p.set_facecolor('#FFA726')
    mean_c = np.mean(confs)
    ax.axvline(mean_c, color=PALETTE['primary'], linestyle='--', linewidth=0.8, alpha=0.7)
    ax.text(0.05, 0.95, f'n = {len(confs)}\nμ = {mean_c:.3f}\nMdn = {np.median(confs):.2f}\nσ = {np.std(confs):.3f}',
            transform=ax.transAxes, fontsize=7, va='top',
            bbox=dict(boxstyle='round,pad=0.4', facecolor=PALETTE['bg_light'], edgecolor=PALETTE['grid'], alpha=0.9))
    ax.set_xlabel('Confidence Score'); ax.set_ylabel('Frequency')
    ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True)); plt.tight_layout()
    return save_fig(fig, output_dir, 'fig06_confidence.png')

def fig07_cause_taxonomy(events, output_dir):
    """Fig 7: Rebellion Cause Taxonomy (2-panel)."""
    causes = Counter(e['_ridda_cause'] for e in events)
    cm = defaultdict(Counter)
    for e in events: cm[e['_ridda_cause']][e['incorporation_mode']] += 1
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(DOUBLE_COL, 3.5), gridspec_kw={'width_ratios': [1, 1.5]})
    # Panel A
    vals = [causes.get(c,0) for c in CAUSE_ORDER]
    ax1.barh(range(len(CAUSE_ORDER)), vals, color=[CAUSE_COLORS[c] for c in CAUSE_ORDER], edgecolor='white', linewidth=0.5, height=0.6)
    ax1.set_yticks(range(len(CAUSE_ORDER))); ax1.set_yticklabels([CAUSE_LABELS[c] for c in CAUSE_ORDER], fontsize=7)
    for i,(bar_v) in enumerate(vals):
        ax1.text(bar_v+0.5, i, f'{bar_v} ({bar_v/len(events)*100:.0f}%)', va='center', fontsize=6.5, color=PALETTE['secondary'])
    ax1.set_xlabel('Number of Events'); ax1.set_title('(a) Rebellion Cause Distribution', fontsize=8, pad=8)
    ax1.xaxis.set_major_locator(ticker.MaxNLocator(integer=True))
    # Panel B
    x = np.arange(len(CAUSE_ORDER)); w = 0.22
    for i,mode in enumerate(MODE_ORDER):
        v = [cm[c][mode] for c in CAUSE_ORDER]
        bars = ax2.bar(x+(i-1)*w, v, w, label=MODE_LABELS[mode], color=MODE_COLORS[mode], edgecolor='white', linewidth=0.3)
        for j,val in enumerate(v):
            if val > 0: ax2.text(x[j]+(i-1)*w, val+0.3, str(val), ha='center', fontsize=6, color=PALETTE['secondary'])
    ax2.set_xticks(x); ax2.set_xticklabels([CAUSE_LABELS_SHORT[c] for c in CAUSE_ORDER], fontsize=6.5, rotation=15, ha='right')
    ax2.set_ylabel('Number of Events'); ax2.set_title('(b) Cause × Incorporation Mode', fontsize=8, pad=8)
    ax2.legend(loc='upper right', frameon=True, edgecolor=PALETTE['grid'], fontsize=6)
    ax2.yaxis.set_major_locator(ticker.MaxNLocator(integer=True)); plt.tight_layout()
    return save_fig(fig, output_dir, 'fig07_cause_taxonomy.png')

def fig08_source_bias(events, output_dir):
    """Fig 8: Source Bias Statistical Test (2-panel)."""
    sm = defaultdict(Counter)
    for e in events: sm[e['_source']][e['incorporation_mode']] += 1
    table = np.array([[sm['tabari'][m] for m in MODE_ORDER],[sm['baladhuri'][m] for m in MODE_ORDER]])
    chi2, p_chi2, dof, expected = scipy_stats.chi2_contingency(table)
    n = table.sum(); cramers_v = np.sqrt(chi2/(n*(min(table.shape)-1)))
    # Fisher pairwise
    fisher = {}
    for mode in MODE_ORDER:
        t2x2 = np.array([[sm['tabari'][mode],60-sm['tabari'][mode]],[sm['baladhuri'][mode],31-sm['baladhuri'][mode]]])
        odds, p_f = scipy_stats.fisher_exact(t2x2)
        fisher[mode] = {'or': odds, 'p': p_f}

    report = {'chi_square': {'statistic': float(chi2), 'p_value': float(p_chi2), 'dof': int(dof)},
              'cramers_v': float(cramers_v), 'expected': expected.tolist(), 'fisher': {m: {'or': float(v['or']), 'p': float(v['p'])} for m,v in fisher.items()}}

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(DOUBLE_COL, 3.2))
    # Panel A: Proportional
    for si,(src,lab,col) in enumerate([('tabari','al-Ṭabarī (n=60)',PALETTE['tabari']),('baladhuri','al-Balādhurī (n=31)',PALETTE['baladhuri'])]):
        tot = sum(sm[src].values())
        props = [sm[src][m]/tot*100 for m in MODE_ORDER]
        x = np.arange(len(MODE_ORDER)); w = 0.35; off = (si-0.5)*w
        bars = ax1.bar(x+off, props, w, label=lab, color=col, edgecolor='white', linewidth=0.3, alpha=0.85)
        for bar, pr, ct in zip(bars, props, [sm[src][m] for m in MODE_ORDER]):
            ax1.text(bar.get_x()+bar.get_width()/2, bar.get_height()+1, f'{pr:.0f}%\n(n={ct})', ha='center', fontsize=5.5, color=PALETTE['secondary'])
    # Significance stars
    for i,mode in enumerate(MODE_ORDER):
        p = fisher[mode]['p']
        star = '***' if p<0.001 else '**' if p<0.01 else '*' if p<0.05 else ''
        if star:
            max_h = max(sm['tabari'][mode]/60, sm['baladhuri'][mode]/31)*100
            ax1.text(i, max_h+10, star, ha='center', fontsize=8, fontweight='bold')
    ax1.set_xticks(range(len(MODE_ORDER))); ax1.set_xticklabels([MODE_LABELS[m] for m in MODE_ORDER], fontsize=7)
    ax1.set_ylabel('Percentage (%)'); ax1.set_title('(a) Mode Distribution by Source', fontsize=8, pad=8)
    ax1.legend(loc='upper left', frameon=True, edgecolor=PALETTE['grid'], fontsize=6); ax1.set_ylim(0,100)
    # Panel B: Residuals
    residuals = (np.array([[sm[s][m] for m in MODE_ORDER] for s in ['tabari','baladhuri']], dtype=float) - expected) / np.sqrt(expected)
    for si,(lab,col) in enumerate([('al-Ṭabarī',PALETTE['tabari']),('al-Balādhurī',PALETTE['baladhuri'])]):
        ax2.bar(np.arange(len(MODE_ORDER))+(si-0.5)*0.35, residuals[si], 0.35, label=lab, color=col, edgecolor='white', alpha=0.85)
    ax2.axhline(0,color=PALETTE['primary'],linewidth=0.5); ax2.axhline(1.96,color='red',linewidth=0.5,linestyle='--',alpha=0.5); ax2.axhline(-1.96,color='red',linewidth=0.5,linestyle='--',alpha=0.5)
    ax2.set_xticks(range(len(MODE_ORDER))); ax2.set_xticklabels([MODE_LABELS[m] for m in MODE_ORDER], fontsize=7)
    ax2.set_ylabel('Standardized Residual'); ax2.set_title('(b) Standardized Residuals', fontsize=8, pad=8)
    ax2.legend(loc='upper right', frameon=True, edgecolor=PALETTE['grid'], fontsize=6)
    sig = 'p < .001' if p_chi2<0.001 else f'p = {p_chi2:.3f}'
    ax2.text(0.02,0.02,f"χ²({dof}) = {chi2:.2f}, {sig}\nCramér's V = {cramers_v:.3f}",transform=ax2.transAxes,fontsize=6,
             bbox=dict(boxstyle='round,pad=0.3',facecolor=PALETTE['bg_light'],edgecolor=PALETTE['grid'],alpha=0.9),va='bottom')
    plt.tight_layout()
    # Save stats
    with open(os.path.join(output_dir, 'source_bias_stats.json'), 'w') as f:
        json.dump(report, f, indent=2)
    return save_fig(fig, output_dir, 'fig08_source_bias.png')

def fig09_term_heatmap(events, output_dir):
    """Fig 9: Evidence Term Heatmap (term × mode)."""
    CATS = {
        'qitāl (fighting)': ['قتال','قاتل','حرب','غزا','حارب'],
        'qatl (killing)': ['قتل','قتلوا','القتل','سيف'],
        'hazīma (defeat)': ['هزم','هزيمة','ظفر','غلب','كسر','انتصر'],
        'ghanīma (spoils)': ['غنم','سبي','سبى','غنيمة','فتح'],
        'ṭāʿa (obedience)': ['طاعة','أطاع','أطاعوا'],
        'tawba (repentance)': ['تاب','رجع','أسلم','أسلموا'],
        'bayʿa (pledge)': ['بايع','بايعوا','البيعة'],
        'ṣulḥ (peace)': ['صالح','صلح','أمان','سلم'],
        'ridda (apostasy)': ['ردة','ارتد','مرتد','ارتدوا'],
        'zakāt': ['الزكاة','زكاة','صدقة','الصدقات'],
    }
    mt = defaultdict(lambda: defaultdict(int)); mc = Counter(e['incorporation_mode'] for e in events)
    for e in events:
        ev = e.get('evidence',[]); txt = ' '.join(ev) if isinstance(ev,list) else str(ev)
        m = e.get('incorporation_mode','')
        for cat,terms in CATS.items():
            if any(t in txt for t in terms): mt[m][cat] += 1
    cats = list(CATS.keys())
    mx = np.zeros((len(cats), len(MODE_ORDER)))
    for i,cat in enumerate(cats):
        for j,mode in enumerate(MODE_ORDER):
            mx[i,j] = (mt[mode].get(cat,0)/mc[mode]*100) if mc[mode] > 0 else 0
    fig, ax = plt.subplots(figsize=(SINGLE_COL+1.5, 4.5))
    cmap = LinearSegmentedColormap.from_list('term', ['#FFFFFF','#FFECB3','#FF8F00','#E65100'])
    im = ax.imshow(mx, cmap=cmap, aspect='auto', vmin=0, vmax=100)
    ax.set_xticks(range(len(MODE_ORDER))); ax.set_xticklabels([MODE_LABELS[m] for m in MODE_ORDER], fontsize=7)
    ax.set_yticks(range(len(cats))); ax.set_yticklabels(cats, fontsize=7)
    for i in range(len(cats)):
        for j in range(len(MODE_ORDER)):
            v = mx[i,j]
            if v > 0: ax.text(j,i,f'{v:.0f}%',ha='center',va='center',fontsize=6.5,fontweight='bold',color='white' if v>50 else PALETTE['primary'])
    ax.axhline(3.5, color=PALETTE['primary'], linewidth=1.0)
    cbar = fig.colorbar(im, ax=ax, shrink=0.8, pad=0.02); cbar.set_label('% Events with Term', fontsize=7)
    plt.tight_layout()
    return save_fig(fig, output_dir, 'fig09_term_heatmap.png')

def fig10_crossmodel_heatmap(output_dir):
    """Fig 10: Cross-Model Agreement Heatmap (4×4 kappa matrix)."""
    models = ['Claude\n(extraction)', 'Gemini\n3.1 Pro', 'Qwen3\n30B', 'ChatGPT\n5.4 Pro']
    # Pre-computed kappa values from 4-way analysis
    kappa_mode = np.array([
        [1.000, 0.662, 0.807, 0.763],
        [0.662, 1.000, 0.658, 0.891],
        [0.807, 0.658, 1.000, 0.757],
        [0.763, 0.891, 0.757, 1.000],
    ])
    kappa_cause = np.array([
        [1.000, 0.436, 0.293, 0.459],
        [0.436, 1.000, 0.794, 0.853],
        [0.293, 0.794, 1.000, 0.706],
        [0.459, 0.853, 0.706, 1.000],
    ])
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(DOUBLE_COL, 3.2))
    cmap = LinearSegmentedColormap.from_list('kappa', ['#FFCDD2','#FFECB3','#C8E6C9','#1B5E20'])

    for ax, data, title in [(ax1, kappa_mode, '(a) Incorporation Mode'), (ax2, kappa_cause, '(b) Rebellion Cause')]:
        im = ax.imshow(data, cmap=cmap, vmin=0.2, vmax=1.0, aspect='equal')
        ax.set_xticks(range(4)); ax.set_xticklabels(models, fontsize=6)
        ax.set_yticks(range(4)); ax.set_yticklabels(models, fontsize=6)
        for i in range(4):
            for j in range(4):
                v = data[i,j]
                if i == j:
                    ax.text(j,i,'—',ha='center',va='center',fontsize=8,color=PALETTE['secondary'])
                else:
                    color = 'white' if v >= 0.7 else PALETTE['primary']
                    ax.text(j,i,f'{v:.3f}',ha='center',va='center',fontsize=7,fontweight='bold',color=color)
        ax.set_title(title, fontsize=8, pad=8)
        # Remove spines for this one
        for spine in ax.spines.values(): spine.set_visible(True); spine.set_linewidth(0.3)
        ax.spines['top'].set_visible(True); ax.spines['right'].set_visible(True)

    cbar = fig.colorbar(im, ax=[ax1,ax2], shrink=0.8, pad=0.02)
    cbar.set_label("Cohen's κ", fontsize=7)
    plt.tight_layout()
    return save_fig(fig, output_dir, 'fig10_crossmodel_heatmap.png')

def fig_s1_mode_pie(events, output_dir):
    """Fig S1: Overall Mode Pie Chart (supplementary)."""
    modes = Counter(e['incorporation_mode'] for e in events if e['incorporation_mode'] in MODE_ORDER)
    fig, ax = plt.subplots(figsize=(SINGLE_COL, SINGLE_COL))
    sizes = [modes[m] for m in MODE_ORDER]; colors = [MODE_COLORS[m] for m in MODE_ORDER]
    labels = [f"{MODE_LABELS[m]}\n({modes[m]})" for m in MODE_ORDER]
    wedges, texts, autotexts = ax.pie(sizes, labels=labels, colors=colors, autopct='%1.0f%%', startangle=90, pctdistance=0.6, wedgeprops=dict(edgecolor='white',linewidth=1.5))
    for t in texts: t.set_fontsize(7)
    for t in autotexts: t.set_fontsize(7); t.set_fontweight('bold'); t.set_color('white')
    ax.set_title('Incorporation Mode Distribution (n=91)', fontsize=8, pad=10); plt.tight_layout()
    return save_fig(fig, output_dir, 'fig_s1_mode_pie.png')


# ============================================================================
# TABLES
# ============================================================================

def write_csv(path, rows):
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader(); writer.writerows(rows)
    print(f"  ✓ {os.path.basename(path)}")

def table01_campaign_phases(scholarly, output_dir):
    phases = scholarly['campaign_phases']['phases']
    rows = [{'Phase': pid.replace('phase_','').upper(), 'Campaign': p['name'],
             'Name_Arabic': p.get('name_ar',''), 'Date': p['date'], 'Duration': p.get('duration','—')}
            for pid, p in sorted(phases.items())]
    write_csv(os.path.join(output_dir, 'table01_campaign_phases.csv'), rows)

def table02_false_prophets(scholarly, output_dir):
    fps = scholarly['biographical_profiles']['false_prophets']
    rows = [{'Name': fp['name_ei3'], 'Arabic': fp['name_ar'], 'Tribe': fp.get('tribe',''),
             'Region': fp.get('region',''), 'Claim': fp.get('claim','')[:80], 'Outcome': fp.get('death','')[:80]}
            for key, fp in fps.items()]
    write_csv(os.path.join(output_dir, 'table02_false_prophets.csv'), rows)

def table03_arabic_indicators(output_dir):
    rows = [
        {'Mode':'SUBJUGATION','Transliteration':'qitāl / ikhḍāʿ','Arabic':'قتال / إخضاع',
         'Indicators':'قتال، قاتل، حرب، غزا، سيف، قتل، هزم، ظفر، غنم، سبي، فتح، حصار',
         'Gloss':'fighting, warred, attacked, raided, sword, killed, defeated, triumphed, plundered, captured, conquered, siege'},
        {'Mode':'SUBMISSION','Transliteration':'ṭāʿa','Arabic':'طاعة',
         'Indicators':'طاعة، أطاع، رجع، تاب، أسلم، بايع، صالح، أمان، سلم، أدى الزكاة',
         'Gloss':'obedience, obeyed, returned, repented, submitted, pledged allegiance, peace, safe conduct, surrendered, paid zakāt'},
        {'Mode':'MIXED','Transliteration':'mukhtalaṭ','Arabic':'مختلط',
         'Indicators':'Combination of qitāl + ṭāʿa indicators','Gloss':'Both military and diplomatic elements'},
    ]
    write_csv(os.path.join(output_dir, 'table03_arabic_indicators.csv'), rows)

def table04_summary_stats(events, output_dir):
    modes = Counter(e['incorporation_mode'] for e in events)
    sources = Counter(e['_source'] for e in events)
    confs = [e['confidence'] for e in events]
    rows = [
        {'Category':'Total events','Value':str(len(events))},
        {'Category':'Unique tribes','Value':str(len(set(e.get('_tribe_normalized') or e['tribe_english'] for e in events)))},
        {'Category':'By Source: al-Ṭabarī','Value':str(sources.get('tabari',0))},
        {'Category':'By Source: al-Balādhurī','Value':str(sources.get('baladhuri',0))},
        {'Category':'SUBJUGATION','Value':str(modes.get('SUBJUGATION',0))},
        {'Category':'SUBMISSION','Value':str(modes.get('SUBMISSION',0))},
        {'Category':'MIXED','Value':str(modes.get('MIXED',0))},
        {'Category':'Year 11 AH','Value':str(sum(1 for e in events if e['year_ah']==11))},
        {'Category':'Year 12 AH','Value':str(sum(1 for e in events if e['year_ah']==12))},
        {'Category':'Confidence (mean)','Value':f'{np.mean(confs):.3f}'},
        {'Category':'Confidence (range)','Value':f'{min(confs):.2f}–{max(confs):.2f}'},
    ]
    write_csv(os.path.join(output_dir, 'table04_summary_stats.csv'), rows)

def table05_commanders(events, output_dir):
    cd = defaultdict(lambda: {'regions': set(), 'modes': Counter(), 'total': 0})
    for e in events:
        c = e.get('_commander_normalized') or e.get('commander_english','')
        if not c or c in ('None','not specified'): continue
        cd[c]['regions'].add(get_region_short(e)); cd[c]['modes'][e.get('incorporation_mode','')] += 1; cd[c]['total'] += 1
    rows = [{'Commander':c,'Events':d['total'],'Regions':','.join(sorted(d['regions'])),
             'SUBJUGATION':d['modes'].get('SUBJUGATION',0),'SUBMISSION':d['modes'].get('SUBMISSION',0),'MIXED':d['modes'].get('MIXED',0)}
            for c,d in sorted(cd.items(), key=lambda x: x[1]['total'], reverse=True)[:8]]
    write_csv(os.path.join(output_dir, 'table05_commanders.csv'), rows)

def table06_cause_mode(events, output_dir):
    cm = defaultdict(Counter)
    for e in events: cm[e['_ridda_cause']][e['incorporation_mode']] += 1
    rows = [{'Cause':CAUSE_LABELS_SHORT[c],'SUBJUGATION':cm[c].get('SUBJUGATION',0),
             'MIXED':cm[c].get('MIXED',0),'SUBMISSION':cm[c].get('SUBMISSION',0),
             'Total':sum(cm[c].values())} for c in CAUSE_ORDER]
    rows.append({'Cause':'Total','SUBJUGATION':sum(r['SUBJUGATION'] for r in rows),
                 'MIXED':sum(r['MIXED'] for r in rows),'SUBMISSION':sum(r['SUBMISSION'] for r in rows),'Total':len(events)})
    write_csv(os.path.join(output_dir, 'table06_cause_mode.csv'), rows)

def table07_validation_summary(events, output_dir):
    rows = [
        {'Metric':'Dataset Size','Value':'91','Notes':''},
        {'Metric':'Geocoding Coverage','Value':'90/91 (98.9%)','Notes':'1 unmapped'},
        {'Metric':'Mean Confidence','Value':f'{np.mean([e["confidence"] for e in events]):.3f}','Notes':''},
        {'Metric':'','Value':'','Notes':''},
        {'Metric':'Cross-Model (Mode)','Value':'','Notes':''},
        {'Metric':"  Fleiss' κ (4 models)",'Value':'0.755','Notes':'Substantial'},
        {'Metric':'  Unanimous (4/4)','Value':'24/30 (80%)','Notes':''},
        {'Metric':'  Consensus (≥3/4)','Value':'26/30 (87%)','Notes':''},
        {'Metric':'','Value':'','Notes':''},
        {'Metric':'Cross-Model (Cause)','Value':'','Notes':''},
        {'Metric':"  Fleiss' κ (4 models)",'Value':'0.566','Notes':'Moderate'},
    ]
    write_csv(os.path.join(output_dir, 'table07_validation_summary.csv'), rows)

def table08_crossmodel_mode(output_dir):
    rows = [
        {'Comparison':'Claude ↔ Gemini','κ':'0.662','Po':'80.0%','Interpretation':'Substantial'},
        {'Comparison':'Claude ↔ Qwen','κ':'0.807','Po':'90.0%','Interpretation':'Substantial'},
        {'Comparison':'Claude ↔ GPT','κ':'0.763','Po':'86.7%','Interpretation':'Substantial'},
        {'Comparison':'Gemini ↔ Qwen','κ':'0.658','Po':'80.0%','Interpretation':'Substantial'},
        {'Comparison':'Gemini ↔ GPT','κ':'0.891','Po':'93.3%','Interpretation':'Almost Perfect'},
        {'Comparison':'Qwen ↔ GPT','κ':'0.757','Po':'86.7%','Interpretation':'Substantial'},
        {'Comparison':"Fleiss' κ (all 4)",'κ':'0.755','Po':'','Interpretation':'Substantial'},
    ]
    write_csv(os.path.join(output_dir, 'table08_crossmodel_mode.csv'), rows)

def table09_crossmodel_cause(output_dir):
    rows = [
        {'Comparison':'Claude ↔ Gemini','κ':'0.436','Po':'56.7%','Interpretation':'Moderate'},
        {'Comparison':'Claude ↔ Qwen','κ':'0.293','Po':'43.3%','Interpretation':'Fair'},
        {'Comparison':'Claude ↔ GPT','κ':'0.459','Po':'60.0%','Interpretation':'Moderate'},
        {'Comparison':'Gemini ↔ Qwen','κ':'0.794','Po':'86.7%','Interpretation':'Substantial'},
        {'Comparison':'Gemini ↔ GPT','κ':'0.853','Po':'90.0%','Interpretation':'Almost Perfect'},
        {'Comparison':'Qwen ↔ GPT','κ':'0.706','Po':'80.0%','Interpretation':'Substantial'},
        {'Comparison':"Fleiss' κ (all 4)",'κ':'0.566','Po':'','Interpretation':'Moderate'},
    ]
    write_csv(os.path.join(output_dir, 'table09_crossmodel_cause.csv'), rows)


# ============================================================================
# ANNOTATION SHEET (Excel)
# ============================================================================

def generate_annotation_excel(events, output_dir):
    """Generate IAA annotation Excel with 4 LLM results + human placeholder."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ⚠ openpyxl not available, skipping Excel generation")
        return None

    random.seed(42)
    mode_ev = defaultdict(list)
    for e in events: mode_ev[e['incorporation_mode']].append(e)
    sample = []
    for mode in MODE_ORDER:
        n = max(3, min(round(len(mode_ev[mode])/len(events)*30), len(mode_ev[mode])))
        sample.extend(random.sample(mode_ev[mode], n))
    if len(sample) > 30: sample = sample[:30]
    random.shuffle(sample)

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Cross-Model Validation"
    headers = ['Event_ID','Source','Tribe_AR','Tribe_EN','Region','Year','Commander','Rebel_Leader','Evidence_AR','Notes',
               '','Claude_Mode','Claude_Cause','','Gemini_Mode','Gemini_Cause',
               '','Qwen_Mode','Qwen_Cause','','GPT_Mode','GPT_Cause',
               '','Human_Mode','Human_Cause','Human_Conf']
    hfill = PatternFill(start_color='D5E8F0',end_color='D5E8F0',fill_type='solid')
    brd = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

    for c,h in enumerate(headers,1):
        cell = ws.cell(row=1,column=c,value=h); cell.font = Font(bold=True,size=8); cell.border = brd; cell.fill = hfill

    for i,e in enumerate(sample):
        row = i+2
        ev_text = ' | '.join(e.get('evidence',[])) if isinstance(e.get('evidence'),list) else str(e.get('evidence',''))
        data = [e.get('_event_id',''),e.get('_source_name',e.get('_source','')),e.get('tribe_arabic',''),
                e.get('tribe_english',''),e.get('region_english',''),e.get('year_ah',''),
                e.get('commander_english','') or '',e.get('rebel_leader_english','') or '',ev_text,e.get('notes','') or '',
                '',e['incorporation_mode'],e.get('_ridda_cause',''),'','','','','','','','','','','','','']
        for c,v in enumerate(data,1):
            cell = ws.cell(row=row,column=c,value=v); cell.border = brd; cell.alignment = Alignment(wrap_text=True,vertical='top'); cell.font = Font(size=8)

    widths = [7,8,12,16,12,5,18,18,45,25,2,13,13,2,13,13,2,13,13,2,13,13,2,13,13,8]
    for i,w in enumerate(widths[:len(headers)],1): ws.column_dimensions[get_column_letter(i)].width = w

    path = os.path.join(output_dir, 'annotation_sheet_IAA_v2.xlsx')
    wb.save(path)
    print(f"  ✓ annotation_sheet_IAA_v2.xlsx (30 events)")
    return path


# ============================================================================
# MAIN
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description='Ridda Wars Unified Figure & Table Generator v3.0')
    parser.add_argument('--data-dir', default='data', help='Directory with enriched/scholarly JSON')
    parser.add_argument('--output-dir', default='figures', help='Output directory')
    parser.add_argument('--figures', nargs='*', default=None, help='Specific figures: 1 2 3 ... 10 s1')
    parser.add_argument('--tables-only', action='store_true', help='Generate tables only')
    parser.add_argument('--no-annotation', action='store_true', help='Skip Excel generation')
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    print(f"\n{'='*65}")
    print(f"RIDDA WARS — UNIFIED FIGURE & TABLE GENERATOR v3.0")
    print(f"Target: MDPI Religions Special Issue")
    print(f"{'='*65}\n")

    enriched, scholarly = load_data(args.data_dir)
    events = enriched['events']
    events = add_cause_taxonomy(events)
    print(f"  Loaded {len(events)} events\n")

    all_figs = {'1','2','3','4','5','6','7','8','9','10','s1'}
    targets = set(args.figures) if args.figures else all_figs

    if not args.tables_only:
        print("FIGURES:")
        if '1'  in targets: fig01_pipeline(args.output_dir)
        if '2'  in targets: fig02_source_comparison(events, args.output_dir)
        if '3'  in targets: fig03_mode_heatmap(events, args.output_dir)
        if '4'  in targets: fig04_temporal(events, args.output_dir)
        if '5'  in targets: fig05_commanders(events, args.output_dir)
        if '6'  in targets: fig06_confidence(events, args.output_dir)
        if '7'  in targets: fig07_cause_taxonomy(events, args.output_dir)
        if '8'  in targets: fig08_source_bias(events, args.output_dir)
        if '9'  in targets: fig09_term_heatmap(events, args.output_dir)
        if '10' in targets: fig10_crossmodel_heatmap(args.output_dir)
        if 's1' in targets: fig_s1_mode_pie(events, args.output_dir)

    print("\nTABLES:")
    table01_campaign_phases(scholarly, args.output_dir)
    table02_false_prophets(scholarly, args.output_dir)
    table03_arabic_indicators(args.output_dir)
    table04_summary_stats(events, args.output_dir)
    table05_commanders(events, args.output_dir)
    table06_cause_mode(events, args.output_dir)
    table07_validation_summary(events, args.output_dir)
    table08_crossmodel_mode(args.output_dir)
    table09_crossmodel_cause(args.output_dir)

    if not args.no_annotation:
        print("\nANNOTATION:")
        generate_annotation_excel(events, args.output_dir)

    # Save enriched with causes
    enriched['events'] = events
    with open(os.path.join(args.output_dir, 'ridda_enriched_with_causes.json'), 'w', encoding='utf-8') as f:
        json.dump(enriched, f, ensure_ascii=False, indent=2)
    print(f"  ✓ ridda_enriched_with_causes.json")

    print(f"\n{'='*65}")
    print(f"ALL OUTPUTS → {args.output_dir}/")
    print(f"{'='*65}\n")


if __name__ == '__main__':
    main()
